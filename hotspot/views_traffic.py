from django.shortcuts import render
from django.db.models import Q
from datetime import datetime, time, timedelta
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.core.cache import cache
from .models import Radacct, UserRouterAccess
from .traffic_models import TrafficLog, DailyWebsiteStats
from .utils import get_allowed_routers
from .log_parser import reverse_dns_cached, simplify_domain 
import re
import socket

# Excel Imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_log_entry(url, method):
    """
    Parses a Squid/Proxy or Mikrotik log entry to extract meaningful fields.
    Handles 'got query from', 'got answer from', and standard URLs.
    """
    url_str = str(url or '').strip()
    method_str = str(method or '').strip()
    combined = url_str + " " + method_str

    log_type = 'Traffic'
    client_ip = None
    domain = None
    dst_ip = None
    port_name = '-'
    protocol = 'HTTP' # Default
    
    if not url_str and not method_str:
        return {'log_type': log_type, 'client_ip': None, 'domain': None, 'dst_ip': None, 'port_name': '-', 'protocol': '-'}

    # 1. Handle Mikrotik DNS style logs
    if 'got query from' in url_str:
        log_type = 'DNS Query'
        match = re.search(r'from ([\d\.]+):(\d+)', url_str)
        if match:
            client_ip = match.group(1)
            port_name = match.group(2)
            protocol = 'DNS'
    elif 'got answer from' in url_str:
        log_type = 'DNS Answer'
        match = re.search(r'from ([\d\.]+):(\d+)', url_str)
        if match:
            dst_ip = match.group(1)
            port_name = match.group(2)
            protocol = 'DNS'
    elif 'from' in url_str and '#' in url_str:
        # Format: "from 10.235.1.253: #1104899860 google.com. A"
        try:
             parts = url_str.split()
             if len(parts) >= 4:
                 candidate = parts[-2]
                 domain = candidate.rstrip('.')
                 log_type = 'DNS Query'
                 protocol = 'DNS'
                 # Find client IP
                 ip_match = re.search(r'from ([\d\.]+)', url_str)
                 if ip_match: client_ip = ip_match.group(1)
        except:
             pass
    elif 'query:' in url_str:
        # Format: "query: #1177260262 input.frontrics.site 52.74.50.179"
        try:
            parts = url_str.split()
            if len(parts) >= 4:
                domain = parts[2].rstrip('.')
                dst_ip = parts[3]
                log_type = 'DNS Query'
                protocol = 'DNS'
        except:
            pass

    # 2. Firewall Forward pattern (IP:PORT->IP:PORT)
    fw_match = re.search(r'([\d\.]+):(\d+)->([\d\.]+):(\d+)', combined)
    if fw_match:
        if not client_ip: client_ip = fw_match.group(1)
        if not dst_ip: dst_ip = fw_match.group(3)
        if port_name == '-': port_name = fw_match.group(4)
        log_type = 'FW'
        if 'proto UDP' in combined: protocol = 'UDP'
        elif 'proto TCP' in combined: protocol = 'TCP'

    # 3. Handle standard HTTP/HTTPS or CONNECT
    if method_str and 'CONNECT' in method_str.upper():
        log_type = 'HTTPS'
        protocol = 'HTTPS'
        domain = url_str.split(':')[0]
    elif '://' in url_str:
        try:
            parts = url_str.split('://')
            protocol = parts[0].upper()
            base = parts[1]
            domain_part = base.split('/')[0]
            domain = domain_part.split(':')[0]
        except:
            domain = url_str
    
    # Cleanup: If domain looks like an IP, move it to dst_ip
    if domain and re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', str(domain)):
        if not dst_ip: dst_ip = domain
        domain = None

    return {
        'log_type': log_type,
        'client_ip': client_ip, 
        'domain': domain,
        'dst_ip': dst_ip,
        'port_name': port_name,
        'protocol': protocol
    }

def lookup_active_sessions_for_logs(logs_list):
    """
    Returns a dict: { ip: [list of sessions] } ordered by start time descending.
    Bounded by time to ensure fast database execution.
    """
    if not logs_list:
        return {}
    
    source_ips = set()
    min_time = None
    
    for log in logs_list:
        if log.source_ip and log.source_ip.strip():
            source_ips.add(log.source_ip.strip())
            
        parsed = parse_log_entry(log.url, log.method)
        if parsed['client_ip']:
            source_ips.add(parsed['client_ip'])
            
        if min_time is None or (log.log_time and log.log_time < min_time):
            min_time = log.log_time
    
    if not source_ips or not min_time:
        return {}

    from django.utils import timezone
    from datetime import timedelta
    
    # Expand window to handle timezone skew (e.g., UTC vs Local) and fetch active users
    search_start_time = min_time - timedelta(days=7)
    
    # Make naive datetime aware if necessary to prevent DB warnings/crashes
    if timezone.is_naive(search_start_time) and timezone.is_aware(timezone.now()):
        search_start_time = timezone.make_aware(search_start_time)
    elif timezone.is_aware(search_start_time) and timezone.is_naive(timezone.now()):
        search_start_time = timezone.make_naive(search_start_time)

    # Use index-friendly bounded query
    sessions = (
        Radacct.objects.using('default')
        .filter(framedipaddress__in=source_ips)
        .filter(acctstarttime__gte=search_start_time)
        .values('username', 'framedipaddress', 'acctstarttime', 'acctstoptime')
        .order_by('framedipaddress', '-acctstarttime')
    )

    ip_sessions = {}
    for s in sessions:
        ip = s['framedipaddress'].strip() if s['framedipaddress'] else ''
        if not ip: continue
        if ip not in ip_sessions:
            ip_sessions[ip] = []
        ip_sessions[ip].append(s)

    return ip_sessions

def enrich_logs(logs, resolve_dns=False):
    """
    Combines raw logs with parsed data (Domain, IP, User) for display.
    """
    enriched = []
    # Pre-fetch overlapping sessions
    active_sessions = lookup_active_sessions_for_logs(logs)
    
    for log in logs:
        parsed = parse_log_entry(log.url, log.method)
        src_ip = parsed['client_ip'] or log.source_ip
        src_ip = src_ip.strip() if src_ip else ''
        
        user = '-'
        log_t = log.log_time
        
        if src_ip and log_t and src_ip in active_sessions:
            # Helper to normalize timezones for comparison
            def normalize_time(t1, ref):
                from django.utils import timezone
                if not t1 or not ref: return t1
                if timezone.is_aware(t1) and timezone.is_naive(ref): return timezone.make_naive(t1)
                if timezone.is_naive(t1) and timezone.is_aware(ref): return timezone.make_aware(t1, ref.tzinfo)
                return t1

            norm_log_t = normalize_time(log_t, active_sessions[src_ip][0]['acctstarttime'])
            from datetime import timedelta
            skew = timedelta(hours=14) # Massive skew allowance for UTC vs Bangkok (+7) issues
            
            # 1. Try to find exact time overlap
            for s in active_sessions[src_ip]:
                start = normalize_time(s['acctstarttime'], norm_log_t)
                stop = normalize_time(s['acctstoptime'], norm_log_t) if s['acctstoptime'] else None
                
                if start and (start - skew) <= norm_log_t and (stop is None or (stop + skew) >= norm_log_t):
                    user = s['username']
                    break
            
            # 2. Add fallback: if no perfect overlap, just use the most recent login for this IP
            if user == '-':
                user = active_sessions[src_ip][0]['username']
        
        # Handle NO MATCH
        dst = parsed['dst_ip'] or log.destination_ip
        if dst == '**NO MATCH**': dst = '--'
        if not dst: dst = '-'
        
        # DNS Resolution Logic
        domain = parsed['domain']
        if not domain and resolve_dns and dst and dst != '-':
            if re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', str(dst)):
                hostname = reverse_dns_cached(dst)
                if hostname:
                    domain = simplify_domain(hostname)
        
        # Normalize type for template (DNS/FW)
        display_type = 'FW'
        if 'DNS' in (parsed['log_type'] or '').upper():
            display_type = 'DNS'
            
        enriched.append({
            'log': log,
            'id': log.id,
            'time': log.log_time,
            'router': log.nas_ip,
            'src_ip': src_ip,
            'dst_ip': dst,
            'domain': domain or '',
            'url': log.url,
            'method': log.method,
            'protocol': parsed['protocol'],
            'port_name': parsed['port_name'],
            'username': user, 
            'log_type': display_type
        })
    return enriched

@login_required
def export_traffic_excel(request):
    """
    Export traffic logs to Excel (.xlsx) with parsed domain/IP info.
    Filters: q, router, users_only, start_date, end_date.
    """
    search_query = request.GET.get('q', '').strip()
    selected_router = request.GET.get('router', '').strip()
    users_only = request.GET.get('users_only', '') == '1'
    web_only = request.GET.get('web_only', '') == '1'
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()
    
    # Access Control
    allowed_routers = get_allowed_routers(request.user)
    
    logs_queryset = TrafficLog.objects.using('default').all().order_by('-log_time')

    # Date Filtering
    try:
        from django.utils.dateparse import parse_datetime
        if start_date_str:
            # Try parsing as datetime, fallback to date
            start_dt = parse_datetime(start_date_str)
            if not start_dt:
                 # If input is just YYYY-MM-DD
                 try:
                     start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
                     start_dt = timezone.make_aware(start_dt)
                 except: pass # Invalid format
            else:
                 if timezone.is_naive(start_dt):
                      start_dt = timezone.make_aware(start_dt)
            if start_dt:
                logs_queryset = logs_queryset.filter(log_time__gte=start_dt)
            
        if end_date_str:
            end_dt = parse_datetime(end_date_str)
            if not end_dt:
                 try:
                     end_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
                     # End of that day
                     end_dt = end_dt.replace(hour=23, minute=59, second=59)
                     end_dt = timezone.make_aware(end_dt)
                 except: pass
            else:
                 if timezone.is_naive(end_dt):
                      end_dt = timezone.make_aware(end_dt)
            if end_dt:
                logs_queryset = logs_queryset.filter(log_time__lte=end_dt)
            
    except Exception as e:
        print(f"Date parsing error: {e}") # Non-blocking

    if allowed_routers is not None:
        logs_queryset = logs_queryset.filter(nas_ip__in=allowed_routers)
    
    if selected_router:
        # Security check
        if allowed_routers is not None and selected_router not in allowed_routers:
            logs_queryset = logs_queryset.none()
        else:
            logs_queryset = logs_queryset.filter(nas_ip=selected_router)
    
    if search_query:
        logs_queryset = logs_queryset.filter(
            Q(nas_ip__icontains=search_query) |
            Q(source_ip__icontains=search_query) |
            Q(destination_ip__icontains=search_query) |
            Q(url__icontains=search_query) |
            Q(method__icontains=search_query)
        )
        
    # User Only Filter (Match Web View Logic)
    if users_only:
        known_ips = (
            Radacct.objects.using('default')
            .exclude(framedipaddress='')
            .values_list('framedipaddress', flat=True)
            .distinct()
        )
        logs_queryset = logs_queryset.filter(source_ip__in=known_ips)
    
    # Limit to 5000 rows for performance
    logs = list(logs_queryset[:5000])
    
    # Pre-fetch Sessions mapped to time ranges
    active_sessions = lookup_active_sessions_for_logs(logs)

    # Resolve/Enrich for export if web_only is active
    final_logs = []
    for log in logs:
        parsed = parse_log_entry(log.url, log.method)
        domain = parsed['domain']
        dst = parsed['dst_ip'] or log.destination_ip
        
        if web_only and not domain:
             continue
             
        # Resolve username exact map
        client_ip = parsed['client_ip'] or log.source_ip
        client_ip = client_ip.strip() if client_ip else ''
        user = '-'
        log_t = log.log_time
        
        if client_ip and log_t and client_ip in active_sessions:
            # Helper to normalize timezones for comparison
            def normalize_time(t1, ref):
                from django.utils import timezone
                if not t1 or not ref: return t1
                if timezone.is_aware(t1) and timezone.is_naive(ref): return timezone.make_naive(t1)
                if timezone.is_naive(t1) and timezone.is_aware(ref): return timezone.make_aware(t1, ref.tzinfo)
                return t1

            norm_log_t = normalize_time(log_t, active_sessions[client_ip][0]['acctstarttime'])
            from datetime import timedelta
            skew = timedelta(hours=14)
            
            for s in active_sessions[client_ip]:
                start = normalize_time(s['acctstarttime'], norm_log_t)
                stop = normalize_time(s['acctstoptime'], norm_log_t) if s['acctstoptime'] else None
                if start and (start - skew) <= norm_log_t and (stop is None or (stop + skew) >= norm_log_t):
                    user = s['username']
                    break
            
            if user == '-':
                user = active_sessions[client_ip][0]['username']
             
        final_logs.append((log, parsed, domain, user))

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Traffic Logs"
    
    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(bottom=Side(style='thin', color='E2E8F0'))
    
    cell_font = Font(name='Calibri', size=10)
    domain_font = Font(name='Calibri', size=10, bold=True, color='1D4ED8')
    ip_font = Font(name='Consolas', size=10, color='0F766E')
    user_font = Font(name='Calibri', size=10, bold=True, color='D97706') # Amber color for user
    
    # Headers (Added Username)
    headers = ['Type', 'Time', 'Router (NAS)', 'Client IP', 'Username', 'Domain / Website', 'Dest IP', 'Port', 'Protocol']
    
    for col, header in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # Data Rows
    for row_idx, (log, parsed, domain, username) in enumerate(final_logs, 2):
        client_ip = parsed['client_ip'] or log.source_ip
        
        # 1. Type
        ws.cell(row=row_idx, column=1, value=parsed['log_type']).font = cell_font
        
        # 2. Time
        ws.cell(row=row_idx, column=2, value=log.log_time.strftime('%Y-%m-%d %H:%M:%S') if log.log_time else '').font = cell_font
        
        # 3. Router
        ws.cell(row=row_idx, column=3, value=log.nas_ip or '').font = ip_font
        
        # 4. Client IP
        ws.cell(row=row_idx, column=4, value=client_ip).font = ip_font
        
        # 5. Username (New)
        ws.cell(row=row_idx, column=5, value=username).font = user_font
        
        # 6. Domain / Website
        display_domain = domain
        if not display_domain:
             display_domain = log.url if log.url and log.url != '-' else (log.method or '')
             
        c = ws.cell(row=row_idx, column=6, value=display_domain)
        c.font = cell_font # Use normal font
        
        # 7. Dest IP
        display_dst = parsed['dst_ip'] or ''
        if display_dst == '**NO MATCH**' or not display_dst:
            display_dst = '--'
        ws.cell(row=row_idx, column=7, value=display_dst).font = ip_font
        
        # 8. Port
        ws.cell(row=row_idx, column=8, value=parsed['port_name'] or '').font = cell_font
        
        # 9. Protocol
        ws.cell(row=row_idx, column=9, value=parsed['protocol'] or '').font = cell_font
    
    # Auto-size columns
    col_widths = [10, 20, 16, 16, 20, 40, 16, 10, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Response
    from datetime import datetime as dt
    filename = f"traffic_logs_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
@login_required
def traffic_log_report(request):
    """
    Traffic Log Report for Computer Crime Act Compliance:
    1. Input: username, date
    2. Logic:
       - Find sessions in Radacct where (User == username) active on the given date.
       - Get the framedipaddress (IP) for those sessions.
       - Query TrafficLog for logs where (source_ip == IP) and log_time falls within Session Start/Stop.
       - Join the data: Log Time, Username (from input), IP (from Log/Session), Website (from Log).
    """
    search_username = request.GET.get('username', '').strip()
    search_date_str = request.GET.get('date', '')
    
    logs_result = []
    
    if search_username and search_date_str:
        try:
            # Parse Date
            naive_date = datetime.strptime(search_date_str, '%Y-%m-%d').date()
            # Assuming timezone aware settings
            tz = timezone.get_current_timezone()
            midday = datetime.combine(naive_date, time(12, 0)) # Just for timezone reference
            if timezone.is_aware(timezone.now()):
                 midday = timezone.make_aware(midday, tz)

            start_of_day = midday.replace(hour=0, minute=0, second=0, microsecond=0)
            end_of_day = midday.replace(hour=23, minute=59, second=59, microsecond=999999)

            # Step 1: Find Sessions for this User/IP active on this Date
            # Condition: Session started before end of day AND (ended after start of day OR is still active)
            
            # Check if input is an IP address
            import re
            is_ip_search = re.match(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$", search_username)
            
            query_kwargs = {
                'acctstarttime__lte': end_of_day
            }
            
            if is_ip_search:
                query_kwargs['framedipaddress'] = search_username
            else:
                query_kwargs['username'] = search_username
                
            sessions = Radacct.objects.using('default').filter(
                **query_kwargs
            ).filter(
                Q(acctstoptime__gte=start_of_day) | Q(acctstoptime__isnull=True)
            )
            
            # Apply router access control
            allowed_routers = get_allowed_routers(request.user)
            if allowed_routers is not None:
                sessions = sessions.filter(nasipaddress__in=allowed_routers)
            
            for session in sessions:
                ip = session.framedipaddress
                start_time = session.acctstarttime
                stop_time = session.acctstoptime if session.acctstoptime else timezone.now()
                
                # Check Timezone awareness to avoid naive/aware comparison error
                if timezone.is_aware(start_time) and timezone.is_naive(start_of_day):
                    start_of_day = timezone.make_aware(start_of_day, start_time.tzinfo)
                    end_of_day = timezone.make_aware(end_of_day, start_time.tzinfo)
                
                # Constrain the search window to the intersection of [Day] and [Session]
                search_start = max(start_time, start_of_day)
                search_end = min(stop_time, end_of_day)
                
                if search_start > search_end:
                    continue
                    
                # Step 2: Query TrafficLog
                # Note: 'source_ip' in TrafficLog must match 'framedipaddress'
                # DEBUG: Use naive comparison if DB is naive
                
                # Check if IP exists
                if not ip:
                    continue
                    
                # If no logs found with strict match, try relaxed window (e.g. +/- 1 min skew)
                # AND check destination_ip because Mikrotik logs might put client IP there
                traffic_logs = TrafficLog.objects.using('default').filter(
                    Q(source_ip=ip) | Q(destination_ip=ip),
                    log_time__gte=search_start - timedelta(minutes=5),
                    log_time__lte=search_end + timedelta(minutes=5)
                )
                
                if allowed_routers is not None:
                    traffic_logs = traffic_logs.filter(nas_ip__in=allowed_routers)

                traffic_logs = traffic_logs.values('log_time', 'source_ip', 'destination_ip', 'url', 'method')

                for log in traffic_logs:
                    # Determine which field holds the meaningful info (URL vs Message)
                    parsed = parse_log_entry(log['url'], log['method'])
                    website_info = parsed['domain']
                    dst_ip = log['destination_ip']

                    # Resolve DNS for Compliance Report
                    if not website_info and dst_ip:
                         if re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', str(dst_ip)):
                              hostname = reverse_dns_cached(dst_ip)
                              if hostname:
                                   website_info = simplify_domain(hostname)
                    
                    if not website_info:
                        website_info = log['url']
                        if website_info == '0.0.0.0' or website_info == '-':
                            website_info = log['method'] or log['url']
                        if log['method'] and len(log['method']) > 20: 
                            website_info = log['method']

                    logs_result.append({
                        'time': log['log_time'],
                        'username': search_username,
                        'ip': ip,
                        'website': website_info,
                        'destination_ip': dst_ip if dst_ip != ip else log['source_ip']
                    })
                    
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
        except Exception as e:
            messages.error(request, f"Error searching logs: {str(e)}")
            import traceback
            print(traceback.format_exc()) # Print to console for debugging
    
    # Sort results by time descending
    logs_result.sort(key=lambda x: x['time'], reverse=True)
    
    # Validation Message for User
    if search_username and not logs_result:
        messages.info(request, f"No traffic logs found for user '{search_username}' on {search_date_str}. Verified sessions existence? { 'Yes' if 'sessions' in locals() and sessions.exists() else 'No' }.")
    
    return render(request, 'hotspot/traffic_log_report.html', {
        'logs': logs_result,
        'search_username': search_username,
        'search_date': search_date_str
    })


# Helpers



@login_required
def traffic_log_list(request):
    """
    General viewer for Traffic Logs.
    Uses cursor-based pagination for performance on large tables.
    """
    search_query = request.GET.get('q', '').strip()
    resolve_dns = request.GET.get('resolve', '0') == '1'
    selected_router = request.GET.get('router', '').strip()
    users_only = request.GET.get('users_only', '') == '1'
    web_only = request.GET.get('web_only', '') == '1'
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str = request.GET.get('end_date', '').strip()
    before_id = request.GET.get('before', '')  # cursor: show records with id < this
    per_page = 50
    
    # Get allowed routers for this user
    allowed_routers = get_allowed_routers(request.user)
    
    # Get distinct routers for dropdown (Cache to avoid scanning millions of rows)
    # Get routers that have logs (already captured)
    logged_routers = cache.get('traffic_routers_list')
    if logged_routers is None:
        logged_routers = list(
            TrafficLog.objects.using('default')
            .exclude(nas_ip__isnull=True)
            .exclude(nas_ip='')
            .values_list('nas_ip', flat=True)
            .distinct()
        )
        cache.set('traffic_routers_list', logged_routers, 60 * 60) # Cache for 1h
        
    if allowed_routers is not None:
        # For restricted users, show their allowed routers even if they have no logs yet
        routers = sorted(list(set(allowed_routers)))
    else:
        # For superusers, show a union of routers with logs and routers in access control
        defined_routers = list(UserRouterAccess.objects.values_list('router_ip', flat=True).distinct())
        routers = sorted(list(set(logged_routers + defined_routers)))
    
    # Build queryset
    logs_queryset = TrafficLog.objects.using('default').all().order_by('-id')
    
    # Router Access Control
    if allowed_routers is not None:
        logs_queryset = logs_queryset.filter(nas_ip__in=allowed_routers)
    
    # Router filter (User selection)
    if selected_router:
        # Security check: if user restricted, ensure selected router is allowed
        if allowed_routers is not None and selected_router not in allowed_routers:
             logs_queryset = logs_queryset.none()
        else:
             logs_queryset = logs_queryset.filter(nas_ip=selected_router)
    
    if search_query:
        logs_queryset = logs_queryset.filter(
            Q(nas_ip__icontains=search_query) |
            Q(source_ip__icontains=search_query) |
            Q(destination_ip__icontains=search_query) |
            Q(url__icontains=search_query) |
            Q(method__icontains=search_query)
        )

    # Date Filtering
    try:
        from django.utils.dateparse import parse_datetime
        if start_date_str:
            start_dt = parse_datetime(start_date_str)
            if not start_dt:
                 try:
                     start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
                     start_dt = timezone.make_aware(start_dt)
                 except: pass
            else:
                 if timezone.is_naive(start_dt):
                      start_dt = timezone.make_aware(start_dt)
            if start_dt:
                logs_queryset = logs_queryset.filter(log_time__gte=start_dt)
            
        if end_date_str:
            end_dt = parse_datetime(end_date_str)
            if not end_dt:
                 try:
                     end_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
                     end_dt = end_dt.replace(hour=23, minute=59, second=59)
                     end_dt = timezone.make_aware(end_dt)
                 except: pass
            else:
                 if timezone.is_naive(end_dt):
                      end_dt = timezone.make_aware(end_dt)
            if end_dt:
                logs_queryset = logs_queryset.filter(log_time__lte=end_dt)
    except Exception as e:
        print(f"Date parsing error: {e}") 
    
    # Pre-filter: only IPs that exist in radacct (known users)
    if users_only:
        known_ips = cache.get('traffic_known_ips')
        if known_ips is None:
            known_ips = list(
                Radacct.objects.using('default')
                .exclude(framedipaddress='')
                .values_list('framedipaddress', flat=True)
                .distinct()
            )
            cache.set('traffic_known_ips', known_ips, 300) # Cache for 5 mins
            
        logs_queryset = logs_queryset.filter(source_ip__in=known_ips)
    
    # Cursor-based pagination: fetch before this ID
    if before_id:
        try:
            logs_queryset = logs_queryset.filter(id__lt=int(before_id))
        except ValueError:
            pass
    
    # Fetch per_page + 1 to check if there are more
    logs_list = list(logs_queryset[:per_page + 1])
    has_next = len(logs_list) > per_page
    logs_list = logs_list[:per_page]
    
    # Get next cursor (last item's ID)
    next_before_id = logs_list[-1].id if logs_list and has_next else None
    
    # Enrich with parsed data (domain, dst IP, port, protocol)
    enriched_logs = enrich_logs(logs_list, resolve_dns=resolve_dns)
    
    # Filter for Web/URL Only if requested
    if web_only:
        enriched_logs = [item for item in enriched_logs if item['domain'] and item['domain'] != '-']

    return render(request, 'hotspot/traffic_log_list.html', {
        'logs_list': logs_list,
        'enriched_logs': enriched_logs,
        'search_query': search_query,
        'resolve_dns': resolve_dns,
        'routers': routers,
        'selected_router': selected_router,
        'users_only': users_only,
        'web_only': web_only,
        'has_next': has_next,
        'next_before_id': next_before_id,
        'before_id': before_id,
        'start_date': start_date_str,
        'end_date': end_date_str,
    })



@login_required
def top_websites(request):
    """
    Top Websites Summary:
    Aggregates TrafficLog by domain (parsed from url/method) and counts
    unique visits within a configurable date range.
    Filters: start_date, end_date, router, top_n.
    """
    import json
    from django.db.models import Count, Sum

    # --- Params ---
    start_date_str = request.GET.get('start_date', '').strip()
    end_date_str   = request.GET.get('end_date', '').strip()
    selected_router = request.GET.get('router', '').strip()
    try:
        top_n = int(request.GET.get('top_n', 20))
        top_n = max(5, min(top_n, 100))
    except ValueError:
        top_n = 20

    # Default: last 7 days if nothing provided
    now = datetime.now()  # USE_TZ=False → use naive datetime throughout
    if not start_date_str and not end_date_str:
        default_start = now - timedelta(days=7)
        start_date_str = default_start.strftime('%Y-%m-%d')
        end_date_str   = now.strftime('%Y-%m-%d')

    # --- Access Control ---
    allowed_routers = get_allowed_routers(request.user)

    # --- Distinct routers for dropdown ---
    logged_routers = cache.get('traffic_routers_list')
    if logged_routers is None:
        logged_routers = list(
            TrafficLog.objects.using('default')
            .exclude(nas_ip__isnull=True).exclude(nas_ip='')
            .values_list('nas_ip', flat=True).distinct()
        )
        cache.set('traffic_routers_list', logged_routers, 3600)
    if allowed_routers is not None:
        routers = sorted(list(set(allowed_routers)))
    else:
        defined_routers = list(UserRouterAccess.objects.values_list('router_ip', flat=True).distinct())
        routers = sorted(list(set(logged_routers + defined_routers)))

    # --- Build base queryset (ALL logs, not just DNS) ---
    qs = TrafficLog.objects.using('default').all()

    if allowed_routers is not None:
        qs = qs.filter(nas_ip__in=allowed_routers)

    if selected_router:
        if allowed_routers is not None and selected_router not in allowed_routers:
            qs = qs.none()
        else:
            qs = qs.filter(nas_ip=selected_router)

    # --- Date filtering (NAIVE datetimes only — project uses USE_TZ=False) ---
    try:
        if start_date_str:
            start_dt = datetime.strptime(start_date_str[:10], '%Y-%m-%d')
            qs = qs.filter(log_time__gte=start_dt)

        if end_date_str:
            end_dt = datetime.strptime(end_date_str[:10], '%Y-%m-%d')
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            qs = qs.filter(log_time__lte=end_dt)
    except Exception as e:
        print(f"[top_websites] date parse error: {e}")

    # --- Mapping: known CDN/technical domains → human-readable service names ---
    DOMAIN_MAP = {
        # Google & YouTube
        'google.com': 'google.com', 'google.co.th': 'google.com', 
        '1e100.net': 'google.com', 'googleusercontent.com': 'google.com',
        'googleapis.com': 'google.com', 'gstatic.com': 'google.com',
        'youtube.com': 'youtube.com', 'googlevideo.com': 'youtube.com', 
        'ytimg.com': 'youtube.com', 'youtu.be': 'youtube.com',
        # Facebook & Instagram
        'facebook.com': 'facebook.com', 'fbcdn.net': 'facebook.com', 
        'facebook.net': 'facebook.com', 'fbsbx.com': 'facebook.com',
        'fb.me': 'facebook.com', 'fb.com': 'facebook.com', 'fb.gg': 'facebook.com',
        'messenger.com': 'facebook.com',
        'instagram.com': 'instagram.com', 'cdninstagram.com': 'instagram.com',
        'ig.me': 'instagram.com', 'igcdn.com': 'instagram.com',
        # TikTok
        'tiktok.com': 'tiktok.com', 'tiktokcdn.com': 'tiktok.com', 
        'tiktokv.com': 'tiktok.com', 'byteoversea.com': 'tiktok.com',
        'ibytedtos.com': 'tiktok.com', 'musical.ly': 'tiktok.com',
        'tiktokcdn-us.com': 'tiktok.com', 'snssdk.com': 'tiktok.com',
        'amemv.com': 'tiktok.com',
        # LINE
        'line.me': 'line.me', 'line-scdn.net': 'line.me', 
        'line-apps.com': 'line.me', 'naver.jp': 'line.me',
        'linecorp.com': 'line.me', 'line.naver.jp': 'line.me',
        # X / Twitter
        'x.com': 'x.com', 'twitter.com': 'x.com', 't.co': 'x.com', 'twimg.com': 'x.com',
        # E-Commerce
        'shopee.co.th': 'shopee.co.th', 'shopeemobile.com': 'shopee.co.th',
        'lazada.co.th': 'lazada.co.th',
        # Microsoft
        'microsoft.com': 'microsoft.com', 'live.com': 'microsoft.com',
        'office.com': 'microsoft.com', 'windows.net': 'microsoft.com',
        'microsoftonline.com': 'microsoft.com', 'msedge.net': 'microsoft.com',
        # Others
        'netflix.com': 'netflix.com', 'spotify.com': 'spotify.com',
        'apple.com': 'apple.com', 'cloudflare.com': 'cloudflare.com',
    }

    def map_domain(raw_domain):
        """Map a technical CDN domain to a recognizable service name."""
        if not raw_domain:
            return raw_domain
        d = raw_domain.lower().strip()
        if d.startswith('www.'):
            d = d[4:]
        # Direct match
        if d in DOMAIN_MAP:
            return DOMAIN_MAP[d]
        # Check if any known suffix matches
        for tech, human in DOMAIN_MAP.items():
            if d.endswith('.' + tech) or d == tech:
                return human
        return d

    # --- Strategy Selector: Today vs Historical ---
    # Sampling for today (live data), Aggregated table for historical
    today = timezone.now().date()
    start_dt_obj = None
    end_dt_obj = None
    try:
        if start_date_str: start_dt_obj = datetime.strptime(start_date_str[:10], '%Y-%m-%d').date()
        if end_date_str: end_dt_obj = datetime.strptime(end_date_str[:10], '%Y-%m-%d').date()
    except: pass

    is_historical = False
    if start_dt_obj and start_dt_obj < today:
        is_historical = True
        
    domain_stats = {} # Final combined stats
    total_processed_requests = 0

    # 1. HISTORICAL DATA (from DailyWebsiteStats)
    if is_historical:
        # Build queryset for stats table
        stats_qs = DailyWebsiteStats.objects.all()
        if start_dt_obj: stats_qs = stats_qs.filter(date__gte=start_dt_obj)
        if end_dt_obj:   stats_qs = stats_qs.filter(date__lte=end_dt_obj)
        if end_dt_obj and end_dt_obj >= today: stats_qs = stats_qs.filter(date__lt=today) # Exclude today from stats table

        if allowed_routers is not None: stats_qs = stats_qs.filter(nas_ip__in=allowed_routers)
        if selected_router: stats_qs = stats_qs.filter(nas_ip=selected_router)

        # Aggregate from table
        # Aggregate from table
        from django.db.models import Sum
        db_stats = stats_qs.values('domain').annotate(
            visits=Sum('visit_count'),
            users=Sum('unique_users')
        ).order_by('-visits')[:top_n * 2]

        for s in db_stats:
            d = map_domain(s['domain'])
            if d not in domain_stats: domain_stats[d] = {'count': 0, 'users': 0}
            domain_stats[d]['count'] += s['visits']
            domain_stats[d]['users'] += s['users']
            total_processed_requests += s['visits']

    # 2. TODAY'S DATA / LIVE SAMPLING (from TrafficLog)
    is_live_needed = (not end_dt_obj or end_dt_obj >= today)
    if is_live_needed:
        live_qs = qs
        if start_dt_obj and start_dt_obj < today:
            live_qs = live_qs.filter(log_time__gte=datetime.combine(today, time.min))
        
        MAX_SAMPLE = 30000  # ลดจาก 100,000 เพื่อกันหน้าเว็บ Timeout (502 Bad Gateway)
        rows = list(live_qs.values('url', 'method', 'destination_ip', 'source_ip')[:MAX_SAMPLE])
        
        raw_stats = {}
        for row in rows:
            parsed = parse_log_entry(row['url'], row['method'])
            domain = (parsed.get('domain') or '').strip().lower()
            dst_ip = (parsed.get('dst_ip') or row.get('destination_ip', '')).strip()
            
            # ถ้ามี Domain ตรงๆ จาก Log ให้ใช้ตัวนั้นและ Map ทันที
            key = None
            if domain and domain != '-':
                key = map_domain(domain)
            elif dst_ip and dst_ip != '-':
                key = dst_ip
            
            if not key: continue
            
            if key not in raw_stats: 
                raw_stats[key] = {
                    'count': 0, 
                    'users': set(), 
                    'is_ip': re.match(r'^\d{1,3}(\.\d{1,3}){3}$', key)
                }
            raw_stats[key]['count'] += 1
            raw_stats[key]['users'].add(row.get('source_ip'))

        # Resolve live IPs
        sorted_raw = sorted(raw_stats.items(), key=lambda x: x[1]['count'], reverse=True)
        public_ips = [k for k,v in sorted_raw if v['is_ip'] and len(k)>7][:50]
        
        from concurrent.futures import ThreadPoolExecutor, as_completed
        ip_to_domain = {}
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {executor.submit(reverse_dns_cached, ip): ip for ip in public_ips}
            for future in as_completed(futures):
                try:
                    ip = futures[future]
                    hostname = future.result()
                    if hostname: ip_to_domain[ip] = map_domain(simplify_domain(hostname))
                except: pass

        # Merge live into domain_stats
        for key, stats in sorted_raw:
            final_key = ip_to_domain.get(key, key)
            if re.match(r'^\d{1,3}(\.\d{1,3}){3}$', final_key): continue # Skip raw IPs
            mapped = map_domain(final_key)
            if mapped not in domain_stats: domain_stats[mapped] = {'count': 0, 'users': 0}
            domain_stats[mapped]['count'] += stats['count']
            
            if isinstance(stats['users'], set):
                domain_stats[mapped]['users'] += len(stats['users'])
            else:
                domain_stats[mapped]['users'] += stats['users']
            
            total_processed_requests += stats['count']

    # --- Sort & Build Return Slice ---
    sorted_domains = sorted(domain_stats.items(), key=lambda x: x[1]['count'], reverse=True)[:top_n]
    total_requests = total_processed_requests
    
    labels      = [d[0] for d in sorted_domains]
    visit_counts = [d[1]['count'] for d in sorted_domains]
    user_counts  = [d[1]['users'] for d in sorted_domains]
    
    pie_labels = labels[:8]
    pie_data   = visit_counts[:8]
    others_sum = sum(visit_counts[8:])
    if others_sum:
        pie_labels.append('Others')
        pie_data.append(others_sum)

    # Table rows
    table_rows = []
    total_count_sum = sum(visit_counts) or 1
    for rank, (domain, stats) in enumerate(sorted_domains, 1):
        pct = round(stats['count'] / total_count_sum * 100, 1) if total_count_sum > 0 else 0
        table_rows.append({
            'rank': rank,
            'domain': domain,
            'count': stats['count'],
            'users': stats['users'],
            'pct': pct,
        })

    return render(request, 'hotspot/top_websites.html', {
        'table_rows': table_rows,
        'total_requests': total_requests,
        'unique_domains': len(domain_stats),
        'routers': routers,
        'selected_router': selected_router,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'top_n': top_n,
        # JSON for charts
        'chart_labels':      json.dumps(labels),
        'chart_counts':      json.dumps(visit_counts),
        'chart_users':       json.dumps(user_counts),
        'pie_labels':        json.dumps(pie_labels),
        'pie_data':          json.dumps(pie_data),
    })

@login_required
def sync_daily_stats(request):
    """
    Manually trigger aggregation for missing days.
    Checks what's in TrafficLog vs DailyWebsiteStats and fills the gaps up to yesterday.
    """
    from datetime import time, timedelta
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    
    # 1. Find the date range available in TrafficLog
    # (Since logs are cleared every 2 days, we only care about what's actually there)
    first_log = TrafficLog.objects.order_by('log_time').first()
    if not first_log:
        messages.warning(request, "ไม่พบข้อมูล Log ในระบบสำหรับนำมาประมวลผล")
        return HttpResponse('<script>window.location.href=document.referrer;</script>')
    
    log_start_date = first_log.log_time.date()
    
    # We aggregate up to yesterday (we don't aggregate 'today' until it's over)
    target_dates = []
    curr = log_start_date
    while curr <= yesterday:
        # Check if already aggregated
        exists = DailyWebsiteStats.objects.filter(date=curr).exists()
        if not exists:
            target_dates.append(curr)
        curr += timedelta(days=1)
        
    if not target_dates:
        messages.info(request, f"ข้อมูลสถิติจนถึงวันที่ {yesterday} ถูกประมวลผลเรียบร้อยแล้ว ไม่ต้องทำอะไรเพิ่มเติม")
        return HttpResponse('<script>window.location.href=document.referrer;</script>')

    # 2. Re-use the logic from management command
    from django.core.management import call_command
    processed_count = 0
    for d in target_dates:
        try:
            date_str = d.strftime('%Y-%m-%d')
            call_command('aggregate_daily_stats', date=date_str)
            processed_count += 1
        except Exception as e:
            print(f"Error aggregating {d}: {e}")

    messages.success(request, f"ประมวลผลข้อมูลย้อนหลังสำเร็จทั้งหมด {processed_count} วัน (จนถึงวันที่ {yesterday})")
    return HttpResponse('<script>window.location.href=document.referrer;</script>')

